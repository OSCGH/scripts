#! /usr/bin/python3
# -*- coding: utf-8 -*-
import openpyxl
import re
import os
import sys
import chardet
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
import shutil
#input:python genxls.py <src>
#output:res.xlsx
#***note: escape chars may not be available***
#rule
#tag->ttype->sheet->col->row->opt
class curexcelpos:
    cursheet=''
    cover = 'cover'
    sheetpos = {}
    basesheet='base'
def openexcel(file, ro=True):
    excel = openpyxl.load_workbook(file,data_only=False,read_only=ro)
    return excel
def closeexcel(excel):
    ret = None
    if(excel != None):
        excel.close()
    return ret
def getdatestr(date):
    return date
def celltostr(cell):
#d:date,b:bool,s:str,n:num,
    val = cell.__str__()
    return val
def readtab(excel,sheet):
    tab=[]
    st=excel[sheet]
    #for row in st.rows:
    for row in st.values:
        line=[]
        for cell in row:
            value=celltostr(cell)
            line.append(value)
        tab.append(line)
    return tab
def saveexcel(excel,file):
    ret = excel.save(file)
    return ret
def writetab(excel,sheet,tab):
    ret = None
    saveexcel(excel)
    return ret
def writecell(data,dstexcel,sheet,col,row):
    dstexcel[sheet]['%s%s'%(col,row)].value = data
    print('%s %s%s=%s' % (sheet,col,row,data))
    return True
def readcell(dstexcel,sheet,col,row):
    celltype = dstexcel[sheet]['%s%s'%(col,row)].data_type
    data = dstexcel[sheet]['%s%s'%(col,row)].value
    if celltype == 'n' and data == None:
        data=''
    return data
def drawborder(resexcel, skipsheet):
    for i in  resexcel.sheetnames:
        if i in skipsheet:
            print('skip sheet %s' % i)
            next
        else:
            print('draw sheet %s border'% i)
            border = Border(left=Side(border_style='thin',color='000000'),
                right=Side(border_style='thin',color='000000'),
                top=Side(border_style='thin',color='000000'),
                bottom=Side(border_style='thin',color='000000'))
            align = Alignment(horizontal='left',vertical='top',wrap_text=True)
            for row in resexcel[i].rows:
                for cell in row:
                    cell.border = border
                    cell.alignment = align
def extract(src,ruletab,dstexcel):
    pos = curexcelpos()
    pos.cursheet = 'sheet1'
    for i in dstexcel.sheetnames:
        pos.sheetpos[i] = {'curcol':'A','currow':2}
    for i in src:
        line = i.replace('\r','').replace('\n','')
        if(re.match('^\**@.*$',line)):
            print('parse %s' %line)
            for j in ruletab:
                if j[0] in '@tag@':
                    print('skip @tag@')
                    next
                if(j[0] in line):
                    data = line.replace(j[0],'')
                    print('data %s' % data)
                    ttype=j[1]
                    opt=j[5]
                    if ttype == 'write':
                        sheet=j[2]
                        if(sheet in '_'):
                            sheet = pos.cursheet
                        pos.cursheet = sheet
                        col=j[3]
                        if(col in '_'):
                            col = pos.sheetpos[pos.cursheet]['curcol']
                        pos.sheetpos[pos.cursheet]['curcol']=col
                        row=j[4]
                        if(row in '_'):
                            row = pos.sheetpos[pos.cursheet]['currow']
                        pos.sheetpos[pos.cursheet]['currow']=row
                        if(opt=='append'):
                            src = readcell(dstexcel,sheet,col,row)
                            if src != '':
                                data='%s\n%s' % (src,data)
                        elif(opt=='replace'):
                            pass
                        elif(opt=='uniqset'):
                            src = readcell(dstexcel,sheet,col,row)
                            if (not data  in src):
                                data='%s\n%s' % (src,data)
                        elif(opt=='mkid'):
                            data = data.replace(' ','')
                            data = data.replace('(','')
                            data = data.replace(')','')
                            data = data.replace(',','.')
                        writecell(data,dstexcel,sheet,col,row)
                    elif ttype == 'newrow':
                        row=pos.sheetpos[pos.cursheet]['currow']
                        pos.sheetpos[pos.cursheet]['currow']=row+1
                        #pos.sheetpos[pos.cursheet]['curcol']='A'
                    elif ttype == 'selectsheet':
                        if data in dstexcel.sheetnames:
                            pos.cursheet=data
                        else:
                            #ws = dstexcel.create_sheet(data)
                            ws=dstexcel.copy_worksheet(dstexcel[pos.basesheet])
                            ws.title = data
                            pos.sheetpos[data] = {'curcol':'A','currow':1}
                            pos.cursheet=data
                    elif ttype == 'basesheet':
                        pos.basesheet = data
                    elif ttype == 'cover':
                        pos.cover = data
        else:
            print('skip %s' %line)
    skipsheet=[pos.cover,pos.basesheet]
    drawborder(resexcel, skipsheet)
if __name__=="__main__":
     #src='src.txt'
     print(sys.argv)
     if(len(sys.argv) != 2):
        print('argc error')
        exit()
     fsrc=sys.argv[1]
     if(not os.path.isfile(fsrc)):
        print('src error')
        exit()
     frule='rule.xlsx'
     ftmp='template.xlsx'
     fdst='res.xlsx'
     if(not os.path.isfile(fdst)):
        shutil.copy(ftmp,fdst)
     ruleexcel=openexcel(frule)
     resexcel=openexcel(fdst,False)
     if(ruleexcel != None):
        ruletab = readtab(ruleexcel,'rule')
        print(ruletab)
        closeexcel(ruleexcel)
     if(resexcel != None):
        coding='utf-8'
        with open(fsrc,mode='rb') as f:
            cont = f.read()
            coding=chardet.detect(cont)['encoding']
            print('%s coding %s'%(fsrc,coding))
        with open(fsrc, mode='r',encoding=coding) as f:
            print('handle src %s' % fsrc)
            extract(f,ruletab,resexcel)
            saveexcel(resexcel,fdst)
            closeexcel(resexcel)

